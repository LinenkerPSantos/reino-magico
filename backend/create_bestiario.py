"""
Cria/atualiza a tabela `bestiary` no banco de dados a partir de
'_Ferramentas/Bestiario.xlsx' (aba Plan1).

Os textos são transcritos exatamente como estão na planilha, célula a célula,
sem resumos. Execute: python create_bestiario.py
"""

import sqlite3
import os
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    sys.exit("Erro: instale openpyxl com  pip install openpyxl")

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
PROJ_DIR  = os.path.dirname(BASE_DIR)
DB_PATH   = os.path.join(BASE_DIR, 'data', 'rpg.db')
XLSX_PATH = os.path.join(PROJ_DIR, '_Ferramentas', 'Bestiario.xlsx')

# Colunas da planilha (A..AL) na ordem em que aparecem
COLUMNS = [
    'nome', 'anotacao', 'descricao', 'aparencia', 'habitat', 'curiosidades',
    'categoria', 'tipo', 'natureza', 'grau_ameaca', 'tamanho', 'comportamento',
    'itens',
    'for_', 'des', 'vit', 'int_', 'pre',
    'vida', 'defesa', 'mana', 'ataque_basico', 'cd_conjurador', 'atributo_conjurador',
    'resistencia', 'fraqueza',
    'habilidade1', 'habilidade2', 'habilidade3', 'habilidade4', 'habilidade5',
    'atletismo', 'reflexo', 'tenacidade', 'vontade', 'foco', 'luta', 'sobrevivencia',
]

# Nome da criatura na planilha -> nome do arquivo de imagem em _Ferramentas/Bestiais
IMAGE_MAP = {
    'Cervo Élfico':          'Cervos Élficos.png',
    'Javali Gigante':        'Javalis Gigantes.png',
    'Urso-Baleia Branco':    'Urso Baleia Branco.png',
    'Pantera Felina Real':   'Panteres Felinos Real.png',
    'Nagas':                 'Naga.png',
}


def _cell(cell):
    v = cell.value
    if v is None:
        return None
    return str(v).strip()


def load_creatures(path: str) -> list:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Plan1']
    creatures = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        nome = _cell(row[0])
        if not nome:
            continue
        d = {}
        for i, col in enumerate(COLUMNS):
            d[col] = _cell(row[i]) if i < len(row) else None
        d['imagem'] = IMAGE_MAP.get(nome, f'{nome}.png')
        creatures.append(d)
    wb.close()
    return creatures


SCHEMA = """
CREATE TABLE IF NOT EXISTS bestiary (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    nome                TEXT NOT NULL UNIQUE,
    anotacao            TEXT,
    descricao           TEXT,
    aparencia           TEXT,
    habitat             TEXT,
    curiosidades        TEXT,
    categoria           TEXT,
    tipo                TEXT,
    natureza            TEXT,
    grau_ameaca         TEXT,
    tamanho             TEXT,
    comportamento       TEXT,
    itens               TEXT,
    for_                INTEGER,
    des                 INTEGER,
    vit                 INTEGER,
    int_                INTEGER,
    pre                 INTEGER,
    vida                INTEGER,
    defesa              INTEGER,
    mana                INTEGER,
    ataque_basico       TEXT,
    cd_conjurador       INTEGER,
    atributo_conjurador TEXT,
    resistencia         TEXT,
    fraqueza            TEXT,
    habilidade1         TEXT,
    habilidade2         TEXT,
    habilidade3         TEXT,
    habilidade4         TEXT,
    habilidade5         TEXT,
    atletismo           INTEGER,
    reflexo             INTEGER,
    tenacidade          INTEGER,
    vontade             INTEGER,
    foco                INTEGER,
    luta                INTEGER,
    sobrevivencia       INTEGER,
    imagem              TEXT
);
"""


def main():
    if not os.path.exists(XLSX_PATH):
        sys.exit(f"Planilha não encontrada: {XLSX_PATH}")

    print(f"Lendo criaturas de: {XLSX_PATH}")
    creatures = load_creatures(XLSX_PATH)
    print(f"  {len(creatures)} criaturas carregadas.")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("DROP TABLE IF EXISTS bestiary")
    conn.executescript(SCHEMA)

    cols = COLUMNS + ['imagem']
    placeholders = ','.join(':' + c for c in cols)
    conn.executemany(
        f"INSERT INTO bestiary({','.join(cols)}) VALUES({placeholders})",
        creatures,
    )

    conn.commit()
    n = conn.execute("SELECT COUNT(*) FROM bestiary").fetchone()[0]
    conn.close()
    print(f"\n✔ Tabela 'bestiary' criada com {n} criaturas.")


if __name__ == '__main__':
    main()
